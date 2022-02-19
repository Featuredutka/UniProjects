import openpyxl
from openpyxl import Workbook
import random

class Robot():
    def __init__(self) -> None:
        self.purpose = ""
        self.envir = ""
        self.chassis = ""
        self.drive = ""
        self.weight = ""
        self.control_sys = ""


source = Workbook()  # Grab the list of the group
source = openpyxl.load_workbook("krbo.xlsx")
sheet = source.active

participants = list(sheet['B'])
participants = [i.value for i in participants][2:20]
source.close()

source = openpyxl.load_workbook("Robot's probability.xlsx")
sheet = source.active

purpose = list(sheet['C1':'J1'][0])
purpose_prob = list(sheet['C3':'J3'][0])

purpose = [i.value for i in purpose]
purpose_prob = [i.value for i in purpose_prob]

# Probabilities for every type of robot

industrial_prob = list(sheet['C'])
industrial_prob = [i.value for i in industrial_prob][3:26]

med_prob = list(sheet['D'])
med_prob = [i.value for i in med_prob][3:26]

domestic_prob = list(sheet['E'])
domestic_prob = [i.value for i in domestic_prob][3:26]

military_prob = list(sheet['F'])
military_prob = [i.value for i in military_prob][3:26]

transport_prob = list(sheet['G'])
transport_prob = [i.value for i in transport_prob][3:26]

research_prob = list(sheet['H'])
research_prob = [i.value for i in research_prob][3:26]

rescue_prob = list(sheet['I'])
rescue_prob = [i.value for i in rescue_prob][3:26]

sport_prob = list(sheet['J'])
sport_prob = [i.value for i in sport_prob][3:26]

# Labels for every type of choice

envir = list(sheet['B'])[3:10]
envir  = [i.value for i in envir]

chassis = list(sheet['B'])[10:15]
chassis  = [i.value for i in chassis]

drive = list(sheet['B'])[15:19]
drive  = [i.value for i in drive]

weight= list(sheet['B'])[19:23]
weight  = [i.value for i in weight]

control = list(sheet['B'])[23:26]
control  = [i.value for i in control]

source.close()

robot = Robot()
dict = {}  # A dictionary to check unique combinations
i = 0

while i < len(participants):

    robot.purpose = random.choices(purpose, weights=purpose_prob)
    
    match robot.purpose[0]:
        case 'Промышленный':
            robot.envir = random.choices(envir, weights=industrial_prob[:7])
            robot.chassis = random.choices(chassis, weights=industrial_prob[7:12])
            robot.drive = random.choices(drive, weights=industrial_prob[12:16])
            robot.weight = random.choices(weight, weights=industrial_prob[16:20])
            robot.control_sys = random.choices(control, weights=industrial_prob[20:23])

        case 'Медицинский':
            robot.envir = random.choices(envir, weights=med_prob[:7])
            robot.chassis = random.choices(chassis, weights=med_prob[7:12])
            robot.drive = random.choices(drive, weights=med_prob[12:16])
            robot.weight = random.choices(weight, weights=med_prob[16:20])
            robot.control_sys = random.choices(control, weights=med_prob[20:23])

        case 'Бытовой':
            robot.envir = random.choices(envir, weights=domestic_prob[:7])
            robot.chassis = random.choices(chassis, weights=domestic_prob[7:12])
            robot.drive = random.choices(drive, weights=domestic_prob[12:16])
            robot.weight = random.choices(weight, weights=domestic_prob[16:20])
            robot.control_sys = random.choices(control, weights=domestic_prob[20:23])

        case 'Боевой':
            robot.envir = random.choices(envir, weights=military_prob[:7])
            robot.chassis = random.choices(chassis, weights=military_prob[7:12])
            robot.drive = random.choices(drive, weights=military_prob[12:16])
            robot.weight = random.choices(weight, weights=military_prob[16:20])
            robot.control_sys = random.choices(control, weights=military_prob[20:23])

        case 'Транспортный':
            robot.envir = random.choices(envir, weights=transport_prob[:7])
            robot.chassis = random.choices(chassis, weights=transport_prob[7:12])
            robot.drive = random.choices(drive, weights=transport_prob[12:16])
            robot.weight = random.choices(weight, weights=transport_prob[16:20])
            robot.control_sys = random.choices(control, weights=transport_prob[20:23])

        case 'Исследовательский':
            robot.envir = random.choices(envir, weights=research_prob[:7])
            robot.chassis = random.choices(chassis, weights=research_prob[7:12])
            robot.drive = random.choices(drive, weights=research_prob[12:16])
            robot.weight = random.choices(weight, weights=research_prob[16:20])
            robot.control_sys = random.choices(control, weights=research_prob[20:23])

        case 'Спасательно-поисковый':
            robot.envir = random.choices(envir, weights=rescue_prob[:7])
            robot.chassis = random.choices(chassis, weights=rescue_prob[7:12])
            robot.drive = random.choices(drive, weights=rescue_prob[12:16])
            robot.weight = random.choices(weight, weights=rescue_prob[16:20])
            robot.control_sys = random.choices(control, weights=rescue_prob[20:23])

        case 'Спортивный':
            robot.envir = random.choices(envir, weights=sport_prob[:7])
            robot.chassis = random.choices(chassis, weights=sport_prob[7:12])
            robot.drive = random.choices(drive, weights=sport_prob[12:16])
            robot.weight = random.choices(weight, weights=sport_prob[16:20])
            robot.control_sys = random.choices(control, weights=sport_prob[20:23])
    
    # Using dictionary with a str hash to check for unique instances
    checksum = str(*robot.purpose) + ' ' + str(*robot.envir) + ' ' + str(*robot.chassis) + ' ' + str(*robot.drive) + ' ' + str(*robot.weight) + ' ' + str(*robot.control_sys)
   
    if checksum in dict.keys():  # TODO If there's a match - try againg
        i -= 1
    else:  # If there's not - whatever
        dict.update({checksum: True})
    i += 1

answer = Workbook()  # Saving results to a file
answer = openpyxl.load_workbook("krbo.xlsx")
sheet = answer.active

for item, i in zip(dict, range(3, 21)):
    cells = sheet.cell(row=i, column=3)
    cells.value = item

answer.save("krbo.xlsx")
answer.close()